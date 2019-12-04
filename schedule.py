# -*- encoding=utf-8 -*-
# import codecs
from datetime import datetime
import datetime as dt
from openpyxl import load_workbook
import sys
import random
import uniout
# import sys
# reload(sys)
# sys.setdefaultencoding('utf-8')
# import pytz
# tz = pytz.timezone('Asia/Shanghai')

print ("Author: Revo, Wechat: lkfo415579, Version: 2.7")
print ("Expired Date: 2020/3/6")
date_str_init = "2019/10/29"
# print (" 請輸入日期:")
print ("Please input Date example : [2019/9/31]:")
date_str = raw_input()
if date_str == "":
    date_str = date_str_init
date = datetime.strptime(date_str, '%Y/%m/%d')
# print (" 日期:%s, 星期:%d" % (date.isoformat(), date.isoweekday()))
print (" Date:%s, Week day:%d" % (date.isoformat(), date.isoweekday()))


loc = "assets/template2.xlsx"
wb = load_workbook(loc)
ws = wb['3cs']
# access all row

rest_rows = {"2cs": [], "3cs": []}


def get_rows_data(ws, dic_id="AF", bench="AH", label='3cs'):
    rows = ws['U3:U846']
    num_rows = 0
    data = []
    type = ""
    for row in rows:
        cell = row[0]
        # special rows, resting area
        if label == '3cs':
            X_pos = "AI"
        else:
            X_pos = "AE"
        if ws['%s%d' % (X_pos, cell.row)].value == 'X':
            rest_rows[label].append(cell.row)
        # assign current type
        t = ws['A' + str(cell.row)].value
        if t:
            type = t
        if cell.value:
            dic = ws[dic_id + '%d' % cell.row].value
            bench_group = ws[bench + '%d' % cell.row].value
            order_id = cell.value
            # check if one of the entry is empty pops error
            if not (dic and bench_group and order_id):
                print (
                    "Error[%s, Row:%d]: is not fillin all data." %
                    (label, cell.row))
                raw_input()
                sys.exit()
            data.append([type, cell.row, dic, order_id, bench_group])
            num_rows += 1
    return data


data = get_rows_data(ws)
data_2cs = get_rows_data(wb['2cs'], "AB", "AD", "2cs")
# access stuffs


def ability(row, ws):
    res = "A"
    res += "C" if ws['G%d' % row].value else ""
    res += "B" if ws['H%d' % row].value else ""
    res += "P" if ws['I%d' % row].value else ""
    res += "R" if ws['J%d' % row].value else ""
    res += "T" if ws['K%d' % row].value else ""
    # special G
    res += "G" if ws['L%d' %
                     row].value and "G" in ws['L%d' %
                                              row].value else ""
    return res


ws_worker = wb['stuff']
rows = ws_worker['A2:A3000']
stuff_data = {}
for row in rows:
    cell = row[0]
    group = ws_worker['L%d' % cell.row].value
    if cell.value and group:
        id = cell.value
        special = str(ws_worker['T%d' % cell.row].value or '')
        name = ws_worker['B%d' % cell.row].value
        type = ws_worker['D%d' % cell.row].value
        abi = ability(cell.row, ws_worker)
        bench = ws_worker['Q%d' % cell.row].value
        relation = ws_worker['O%d' % cell.row].value
        relation = ":" + str(relation) if relation else ""
        t = [str(id), name, type, abi, group, special, bench, relation]
        if group in stuff_data:
            stuff_data[group].append(t)
        else:
            stuff_data[group] = [t]

# access holiday
ws_worker = wb['holiday']
rows = ws_worker['C2:C2000']
holiday = {}
for row in rows:
    cell = row[0]
    if cell.value:
        try:
            start_str = str(cell.value).replace("-", "")[:8]
            end_str = str(ws_worker['D%d' % cell.row].value).replace("-", "")[:8]
            start_date = datetime.strptime(start_str, '%Y%m%d')
            end_date = datetime.strptime(end_str, '%Y%m%d')
            if date > start_date and date < end_date or date == start_date or date == end_date:
                id = str(ws_worker['B%d' % cell.row].value)
                holiday[id] = [id, start_date, end_date]
        except BaseException as e:
            print ("Holiday Error:", e)
            pass

# generating pattern table
patter_start_date = "2019/08/19"
patter_start_date = datetime.strptime(patter_start_date, '%Y/%m/%d')
table = [["A", "E", "D", "G", "C", "F"],
         ["A", "E", "D", "G", "B", "F"],
         ["A", "E", "C", "G", "B", "F"],
         ["A", "D", "C", "G", "B", "F"],
         ["A", "D", "C", "G", "B", "E"],
         ["A", "D", "C", "F", "B", "E"],
         ["D", "G", "C", "F", "B", "E"],
         ["D", "G", "C", "F", "A", "E"],
         ["D", "G", "B", "F", "A", "E"],
         ["C", "G", "B", "F", "A", "E"],
         ["C", "G", "B", "F", "A", "D"],
         ["C", "G", "B", "E", "A", "D"],
         ["C", "F", "B", "E", "A", "D"],
         ["C", "F", "B", "E", "D", "G"],
         ["C", "F", "A", "E", "D", "G"],
         ["B", "F", "A", "E", "D", "G"],
         ["B", "F", "A", "E", "C", "G"],
         ["B", "F", "A", "D", "C", "G"],
         ["B", "E", "A", "D", "C", "G"],
         ["B", "E", "A", "D", "C", "F"],
         ["B", "E", "D", "G", "C", "F"]]


def retrieve_today(patter_start_date, table):
    table_data = {}
    for i in range(200):
        table_data[patter_start_date.isoformat()] = table[i % len(table)]
        patter_start_date += dt.timedelta(days=1)

    # main function, (type, row_id, dic)
    # stuff[abi]=[id, name, type, abi, group, special]
    # data=[type, cell.row, dic]
    data.sort(key=lambda x: (x[2], x[1]), reverse=True)
    data_2cs.sort(key=lambda x: (x[2], x[1]), reverse=True)
    today = table_data[date.isoformat()]
    return today


today = retrieve_today(patter_start_date, table)
table = [["A", "D", "F", "C", "E", "G"],
         ["A", "D", "F", "B", "E", "G"],
         ["A", "C", "F", "B", "E", "G"],
         ["A", "C", "F", "B", "D", "G"],
         ["A", "C", "E", "B", "D", "G"],
         ["A", "C", "E", "B", "D", "F"],
         ["C", "E", "G", "B", "D", "F"],
         ["C", "E", "G", "A", "D", "F"],
         ["B", "E", "G", "A", "D", "F"],
         ["B", "E", "G", "A", "C", "F"],
         ["B", "D", "G", "A", "C", "F"],
         ["B", "D", "G", "A", "C", "E"],
         ["B", "D", "F", "A", "C", "E"],
         ["B", "D", "F", "C", "E", "G"]]

today_2cs = retrieve_today(patter_start_date, table)

# check if only one type of players exists


def IsOnly(can_stuffs, abi, index):
    for stuff in can_stuffs:
        if abi not in stuff[index]:
            return False
    return True


def filter(can_stuffs, bench_group):
    if bench_group:
        return [x for x in can_stuffs if x[6] == bench_group or not x[6]]


def output_info(stuff):
    # wont display ability's G and A
    stuff[3] = "".join(set(stuff[3])).replace(
        "A", "").replace("I", "").replace("G", "")
    return stuff[0], stuff[1], str(stuff[3] + stuff[5])


def main(i, ii, iii, ci, cii, ciii, ws, data, today, cs='3'):
    for row in data:
        type, row_id, dic, order_id, bench_group = row
        if type == u'海皇神殿':
            type = u'庄荷'
        if type != u'庄荷':
            type = u'傍庄'
        #
        can_stuffs = stuff_data[today[i] + cs][:]
        can_stuffs.extend(stuff_data[today[ii] + cs][:])
        if iii != -1:
            can_stuffs.extend(stuff_data[today[iii] + cs][:])
        # filter those has bench_group stuffs
        can_stuffs = filter(can_stuffs, bench_group)
        # shuffle stuffs
        random.shuffle(can_stuffs)
        # sort by abi first
        can_stuffs.sort(key=lambda x: x[3], reverse=True)
        # remove dayoff stuff
        try:
            stuff = can_stuffs.pop()
        except BaseException:
            print(
                "%cs, row:%d, type:%s is not enough stuff for fillin." %
                (cs, row_id, type))
            stuff = None
        # stuff[abi]=[id, name, type, abi, group, special, bench]
        while stuff:
            # X + ~ + others classes special handling, change forcely into B
            # type
            if stuff[2] != u'庄荷' or ('X' in stuff[5] or '~' in stuff[5]):
                stuff[2] = u'傍庄'
            # not in holiday,
            # dic is matched
            # type is matched
            if stuff[0] not in holiday and dic in stuff[3] and type == stuff[2]:
                # check if it is a new stuff
                if '*' in stuff[5]:
                    first_table_id = unicode(ws['B%d' % row_id].value)
                    if not first_table_id.isdigit(
                    ) or first_table_id[:2] == '23' or first_table_id[:2] == '35':
                        # print("New Player can't be 23 and 35 and str table. reput it to the end of queue")
                        stuff = can_stuffs.pop() if can_stuffs else None
                        continue
                # check if it has special bench
                if stuff[6]:
                    if bench_group != stuff[6]:
                        stuff = can_stuffs.pop() if can_stuffs else None
                        continue

                # normally assign
                stuff_data[stuff[4]].remove(stuff)
                break
            else:
                # reget another stuff
                try:
                    stuff = can_stuffs.pop()
                except BaseException:
                    # print("%cs, 行:%d, 類型:%s 不夠員工填上。" % (cs, row_id, type.encode('utf-8')))
                    print(
                        "%cs, row:%d, type:%s is not enough stuff for fillin." %
                        (cs, row_id, type))
                    stuff = None
                    break
        stuff = [None, None, "", "", "", "", "", ""] if not stuff else stuff
        # on duty record
        # base_char = ord('T') + 4*(i/2)
        ws['%s%d' % (ci, row_id)] = stuff[0]
        ws['%s%d' % (cii, row_id)] = stuff[1]
        # wont display ability's G and A
        stuff[3] = "".join(set(stuff[3])).replace(
            "A", "").replace("I", "").replace("G", "")
        #
        relation = stuff[7]
        ws['%s%d' % (ciii, row_id)] = str(stuff[3] + stuff[5] + relation)
    # refresh date
    ws['U1'] = date_str
    ws['Y1'] = date_str
    ws['AC1'] = date_str

# 3cs


print ("3cs table", today)
main(0, 1, -1, "T", "U", "V", ws, data, today)
main(2, 3, -1, "X", "Y", "Z", ws, data, today)
main(4, 5, -1, "AB", "AC", "AD", ws, data, today)

# 2cs
print ("=" * 100)
ws = wb['2cs']
print ("2cs table", today_2cs)
main(0, 1, 2, "T", "U", "V", ws, data_2cs, today_2cs, '2')
main(3, 4, 5, "X", "Y", "Z", ws, data_2cs, today_2cs, '2')

print ("=" * 100)
# print (" 總共有%d員工今天請假" % len(holiday))
print (" Today: %d Stuffs day-off" % len(holiday)),
print (holiday)
print ("=" * 100)
print ("You have stuffs are not assigned yet", stuff_data)
print ("Putting all resting stuffs into XXX columns....")


def find_holiday_group(today):
    all = ["A", "B", "C", "D", "E", "F", "G"]
    for s in all:
        if s not in today:
            return s
    return ""


def rest(stuff_data, today, cs="3"):
    stuff_data = stuff_data.copy()
    ws = wb['%scs' % cs]
    another_cs = "2" if cs == "3" else "3"
    y_axis = {0: ["T", "U", "V"], 1: ["X", "Y", "Z"], 2: ["AB", "AC", "AD"]}
    #
    counter = 0
    all = []
    regular_week = find_holiday_group(today)
    for type in stuff_data:
        if type == regular_week + cs or type[-1] == another_cs:
            all.append(type)
    for t in all:
        del stuff_data[t]

    for k in range(0, int(cs)):
        can_stuffs = stuff_data[today[k * 2] + cs] + stuff_data[today[k * 2 + 1] + cs]
        r = 0
        for i, stuff in enumerate(can_stuffs):
            if stuff[0] in holiday:
                continue
            try:
                row = rest_rows["%scs" % cs][r]
            except:
                print("Don't not have enough X rows for saving rest of the people!!!!")
                raw_input()

            id, name, last = output_info(stuff)
            ws['%s%d' % (y_axis[k][0], row)] = id
            ws['%s%d' % (y_axis[k][1], row)] = name
            ws['%s%d' % (y_axis[k][2], row)] = last
            r += 1
            counter += 1

    print("Resting stuffs:%d" % counter)


rest(stuff_data, today)
rest(stuff_data, today_2cs, "2")

print (" Writting into EXCEL.....")
try:
    wb.save("output.xlsx")
except Exception as e:
    print(e)
    raw_input()
#
print ("=" * 100)
# disposal table
loc = "assets/disposal.xlsx"
wb = load_workbook(loc)
ws = wb[u'新更表']


def read_stuff():
    ranges = ['I', 'M', 'Q']
    disposal_stuffs = {}
    for range in ranges:
        rows = ws["%s8:%s100" % (range, range)]
        for row in rows:
            cell = row[0]
            if cell.value and (isinstance(cell.value, int) or isinstance(cell.value, long)):
                name = ws["%s%d" % (chr(ord(range) + 1), cell.row)].value
                group = ws["%s%d" % (chr(ord(range) + 2), cell.row)].value
                extra = ws["%s%d" % (chr(ord(range) + 3), cell.row)].value
                stuff = {"id": str(cell.value), "name": name, "group": group, "extra": extra}
                if group not in disposal_stuffs:
                    disposal_stuffs[group] = [stuff]
                else:
                    disposal_stuffs[group].append(stuff)
    return disposal_stuffs


def assign(dis, today, holiday):
    # holiday
    r = 0
    for group in dis:
        for d in dis[group]:
            if d['id'] in holiday:
                ws['Q%d' % (26 + r)] = d['id']
                ws['R%d' % (26 + r)] = d['name']
                ws['S%d' % (26 + r)] = d['group']
                ws['T%d' % (26 + r)] = d['extra']
                r += 1
                dis[group].remove(d)
    #
    pos = [9, 17, 14, 18, 8]
    #
    for II in range(0, 5, 2):
        can_stuffs = dis[today[II]] + dis[today[II + 1]]
        random.shuffle(can_stuffs)
        id_alpha = ord('I') + int(4 * (II / 2))
        for i, p in enumerate(pos):
            if i < len(can_stuffs):
                stuff = can_stuffs[i]
                ws["%s%d" % (chr(id_alpha), pos[i])] = stuff['id']
                ws["%s%d" % (chr(id_alpha + 1), pos[i])] = stuff['name']
                ws["%s%d" % (chr(id_alpha + 2), pos[i])] = stuff['group']
                ws["%s%d" % (chr(id_alpha + 3), pos[i])] = stuff['extra']
            else:
                ws["%s%d" % (chr(id_alpha), pos[i])] = ""
                ws["%s%d" % (chr(id_alpha + 1), pos[i])] = ""
                ws["%s%d" % (chr(id_alpha + 2), pos[i])] = ""
                ws["%s%d" % (chr(id_alpha + 3), pos[i])] = ""
    # regular holiday
    holiday_group = find_holiday_group(today)
    can_stuffs = dis[holiday_group]
    for i, stuff in enumerate(can_stuffs):
        ws["I%d" % (26 + i)] = stuff['id']
        ws["J%d" % (26 + i)] = stuff['name']
        ws["K%d" % (26 + i)] = stuff['group']
        ws["L%d" % (26 + i)] = stuff['extra']

    pass



disposal_stuffs = read_stuff()
assign(disposal_stuffs, today, holiday)
ws['J2'] = date_str
try:
    wb.save("disposal_output.xlsx")
except Exception as e:
    print(e)
    raw_input()
print (" Finished Writting EXCEL.....")
stop = raw_input()
pass
